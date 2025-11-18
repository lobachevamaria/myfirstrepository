import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def csv_to_xlsx(csv_path: str, xlsx_path: str) -> None:
    """
    Преобразует CSV-файл в Excel (XLSX) с проверкой структуры и корректности данных.

    CSV-файл должен содержать строку заголовков и хотя бы одну строку данных.
    Каждая колонка CSV будет перенесена в отдельный столбец листа Excel.

    Аргументы:
        csv_path: Путь к исходному CSV-файлу.
        xlsx_path: Путь к создаваемому XLSX-файлу.


    Ошибки:
        FileNotFoundError: Если CSV-файл отсутствует.
        ValueError: Если CSV не содержит заголовков или пуст.
                Если структура CSV-файла некорректна.
    """

    csv_file = Path(csv_path)
    if not csv_file.exists():
        raise FileNotFoundError("Файл не найден")
    if csv_file.suffix != ".csv":
        raise ValueError("Неверный тип файла")

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    with open(csv_path, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        rows = list(reader)
    if len(rows) == 0:
        raise ValueError("Файл не содержит данных")
    if not reader.fieldnames:
        raise ValueError("Файл не содержит заголовка")

    ws.append(reader.fieldnames)
    r_count = 0
    for row in rows:
        r_count += 1

        data_for_ex = []
        for title in reader.fieldnames:
            data_for_ex.append(row[title])
        ws.append(data_for_ex)
    if r_count == 0:
        raise ValueError("Нет данных")

    for col_index in range(1, len(reader.fieldnames) + 1):
        column_letter = get_column_letter(col_index)
        max_len = 0

        for row in ws[column_letter]:
            if row.value is not None:
                max_len = max(max_len, len(str(row.value)))

        m_width = max(max_len + 2, 8)
        ws.column_dimensions[column_letter].width = m_width

    xlsx_path = Path(xlsx_path)
    wb.save(xlsx_path)
